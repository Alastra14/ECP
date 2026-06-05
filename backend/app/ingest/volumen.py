"""Ingesta de VOLUMEN desde 'VOL PRODUCCIÓN.xlsx'.

Hoja 'COSTOS UNITARIOS PT CM': volumen por código de mercancía y mes (1..4),
agrupado bajo cada FAM (celda combinada). Se agrega por línea × mes.
Volumen del reporte = kgs-Mil, así que dividimos kg entre 1000.
"""
from __future__ import annotations

from pathlib import Path

import openpyxl

from ..domain import linea_desde_grupo, ESCENARIO_REAL
from ..config import settings

HOJA = "COSTOS UNITARIOS PT CM"
FILA_INICIO = 9
COLS_MES = {1: 3, 2: 4, 3: 5, 4: 6}  # C,D,E,F = meses 1..4 del bloque "Volumen de Prod"


def ingerir_volumen(path: str | Path, anio: int | None = None) -> list[dict]:
    anio = anio or settings.anio_activo
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    if HOJA not in wb.sheetnames:
        wb.close()
        return []
    ws = wb[HOJA]

    acum: dict[tuple[str, int], float] = {}
    fam_actual = None
    vacias = 0
    for r in range(FILA_INICIO, ws.max_row + 1):
        fam_cell = ws.cell(row=r, column=1).value  # col A (combinada)
        cod = ws.cell(row=r, column=2).value       # col B
        if fam_cell:
            fam_actual = fam_cell
        if not fam_cell and not cod:
            vacias += 1
            if vacias > 5:
                break
            continue
        vacias = 0
        linea = linea_desde_grupo(fam_actual)
        if not linea:
            continue
        for mes, col in COLS_MES.items():
            v = ws.cell(row=r, column=col).value
            if isinstance(v, (int, float)):
                acum[(linea, mes)] = acum.get((linea, mes), 0.0) + float(v)
    wb.close()

    comps = []
    for (linea, mes), kg in acum.items():
        comps.append({
            "linea": linea,
            "periodo": f"{anio:04d}-{mes:02d}",
            "concepto": "volumen",
            "escenario": ESCENARIO_REAL,
            "monto": round(kg / 1000.0, 3),  # kgs-Mil
        })
    return comps
