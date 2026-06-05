"""Extractor de referencia: lee la hoja '1-ECP MES' de los reportes RESULTADO.

Cada archivo `2604-ECP CM-XXX.xlsx` es una línea. La hoja trae varios meses
(columnas B..E = REAL) y la columna F = PPTO del mes activo. De aquí salen los
componentes EXACTOS, que sirven como seed para la demo y como oráculo de pruebas.
"""
from __future__ import annotations

import datetime as _dt
from pathlib import Path

import openpyxl

from ..domain import FILA_A_CONCEPTO, linea_desde_codigo_ecp, ESCENARIO_REAL, ESCENARIO_PPTO

HOJA = "1-ECP MES"
COL_PPTO = 6  # F


def _periodo(d) -> str | None:
    if isinstance(d, (_dt.datetime, _dt.date)):
        return f"{d.year:04d}-{d.month:02d}"
    return None


def extraer_resultado(path: str | Path) -> list[dict]:
    """Devuelve componentes {linea, periodo, concepto, escenario, monto} de un archivo."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    if HOJA not in wb.sheetnames:
        wb.close()
        return []
    ws = wb[HOJA]

    linea = linea_desde_codigo_ecp(ws.cell(row=3, column=1).value)
    if not linea:
        wb.close()
        return []

    # Mapa columna -> periodo (REAL), leyendo la fila de encabezado (fila 5).
    col_periodo: dict[int, str] = {}
    periodo_activo = None
    for col in range(2, 6):  # B..E
        per = _periodo(ws.cell(row=5, column=col).value)
        if per:
            col_periodo[col] = per
            periodo_activo = per  # la última (col E) es el mes activo

    comps: list[dict] = []
    for fila, concepto in FILA_A_CONCEPTO.items():
        # REAL por cada mes presente
        for col, per in col_periodo.items():
            val = ws.cell(row=fila, column=col).value
            if isinstance(val, (int, float)):
                comps.append({"linea": linea, "periodo": per, "concepto": concepto,
                              "escenario": ESCENARIO_REAL, "monto": float(val)})
        # PPTO (columna F) -> se asocia al mes activo
        if periodo_activo is not None:
            valp = ws.cell(row=fila, column=COL_PPTO).value
            if isinstance(valp, (int, float)):
                comps.append({"linea": linea, "periodo": periodo_activo, "concepto": concepto,
                              "escenario": ESCENARIO_PPTO, "monto": float(valp)})
    wb.close()
    return comps


def extraer_carpeta_resultado(carpeta: str | Path) -> list[dict]:
    """Recorre todos los `*ECP CM-*.xlsx` de la carpeta RESULTADO."""
    carpeta = Path(carpeta)
    comps: list[dict] = []
    for path in sorted(carpeta.glob("*ECP CM-*.xlsx")):
        if path.name.startswith("~$"):
            continue
        comps.extend(extraer_resultado(path))
    return comps
