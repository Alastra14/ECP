"""Ingesta de COSTO DE MATERIA PRIMA desde 'Copia ... MP.xlsx', hoja 'real'.

Columnas: B=mes cierre, C=FAM, D=Commodity, E=Qty_Consumo, F=Costo Total.
Se agrega Costo Total por línea × mes (en miles).
"""
from __future__ import annotations

from pathlib import Path

import openpyxl

from ..domain import linea_desde_grupo, ESCENARIO_REAL
from ..config import settings

HOJA = "real"
FILA_INICIO = 4
COL_MES, COL_FAM, COL_COSTO = 2, 3, 6  # B, C, F


def ingerir_mp(path: str | Path, anio: int | None = None) -> list[dict]:
    anio = anio or settings.anio_activo
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    if HOJA not in wb.sheetnames:
        wb.close()
        return []
    ws = wb[HOJA]

    acum: dict[tuple[str, int], float] = {}
    vacias = 0
    for r in range(FILA_INICIO, ws.max_row + 1):
        mes = ws.cell(row=r, column=COL_MES).value
        fam = ws.cell(row=r, column=COL_FAM).value
        costo = ws.cell(row=r, column=COL_COSTO).value
        if mes is None and fam is None:
            vacias += 1
            if vacias > 5:
                break
            continue
        vacias = 0
        if not isinstance(mes, (int, float)) or not isinstance(costo, (int, float)):
            continue
        linea = linea_desde_grupo(fam)
        if not linea:
            continue
        acum[(linea, int(mes))] = acum.get((linea, int(mes)), 0.0) + float(costo)
    wb.close()

    return [{
        "linea": linea, "periodo": f"{anio:04d}-{mes:02d}",
        "concepto": "mp", "escenario": ESCENARIO_REAL,
        "monto": round(v / 1000.0, 3),  # miles
    } for (linea, mes), v in acum.items()]
